
import pytesseract
from pdf2image import convert_from_path
import re
import regex
import time
import openpyxl
from openpyxl import Workbook
from PIL import ImageOps
import os
import argparse
from pathlib import Path
import sys

# Configurar ruta a tesseract
base_dir = os.path.dirname(os.path.abspath(__file__))
pytesseract.pytesseract.tesseract_cmd = os.path.join(base_dir, "Tesseract-OCR", "tesseract.exe")
os.environ["TESSDATA_PREFIX"] = os.path.join(base_dir, "Tesseract-OCR", "tessdata")

def normalize_number(s: str) -> str:
    last = max(s.rfind(','), s.rfind('.'))
    int_part, dec_part = s[:last], s[last+1:]
    int_part = re.sub(r'[^\d]', '', int_part)
    return f'{int_part},{dec_part}'

def normalize_invoice(s: str) -> str:
    m = re.search(r'0+([1-9]\d*)', s)
    if m:
        return m.group(1)
    out = s.lstrip('0')
    return out if out else '0'

def save_to_excel(valor_factura, valor_tasa, pagina_pdf, filename):
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        next_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'No. Factura'
        ws['B1'] = 'Tasa de Cambio'
        ws['C1'] = 'Página PDF'
        ws['D1'] = 'Archivo Origen'
        next_row = 2

    ws[f'A{next_row}'] = valor_factura
    ws[f'B{next_row}'] = valor_tasa
    ws[f'C{next_row}'] = pagina_pdf[0]
    ws[f'D{next_row}'] = pagina_pdf[1]
    wb.save(filename)
    print(f"Datos agregados en fila {next_row} de {filename}")

def format_elapsed_time(elapsed_seconds):
    hours = int(elapsed_seconds // 3600)
    minutes = int((elapsed_seconds % 3600) // 60)
    seconds = int(elapsed_seconds % 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}" if hours > 0 else f"{minutes:02d}:{seconds:02d}"

def main():
    print(f"🐍 Versión de Python:", sys.version.split()[0])
    parser = argparse.ArgumentParser()
    parser.add_argument("--ruta", required=True, help="Ruta principal donde buscar los archivos PDF")
    parser.add_argument("--dry-run", action="store_true", help="No guarda Excel, solo muestra resultados")
    args = parser.parse_args()

    start_time = time.time()
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    excel_filename = f"resultados_{timestamp}.xlsx"
    output_dir = f"ocr_results_{timestamp}"
    os.makedirs(output_dir, exist_ok=True)

    print(f"🚀 Iniciando procesamiento a las {time.strftime('%H:%M:%S')}")
    print(f"📁 Buscando PDFs en: {args.ruta}")
    print(f"📊 Archivo de salida: {excel_filename}")
    print("-" * 50)

    pdf_files = list(Path(args.ruta).rglob("*Declaracion Importacion*.pdf"))
    print(f"📄 Total de archivos PDF encontrados: {len(pdf_files)}")
    print("-" * 50)

    pat_fuzzy = regex.compile(r'(?i)(procedencia|transporte|bandera|destino){e<=2}', flags=0)
    total_resultados = 0

    for pdf_path in pdf_files:
        print(f"📂 Procesando archivo: {pdf_path}")
        conversion_start = time.time()
        try:
            pages = convert_from_path(pdf_path, poppler_path=os.path.join(base_dir, "poppler", "Library", "bin"))
        except Exception as e:
            print(f"❌ Error al convertir PDF: {e}")
            continue
        conversion_time = time.time() - conversion_start
        print(f"✅ PDF convertido en {conversion_time:.2f} segundos | {len(pages)} páginas")

        for page_num, image in enumerate(pages):
            page_start_time = time.time()
            elapsed_total = time.time() - start_time
            print(f"\n📄 Página {page_num + 1}/{len(pages)} | Tiempo total: {format_elapsed_time(elapsed_total)}")

            gray_image = ImageOps.grayscale(image)
            ocr_start = time.time()
            text = pytesseract.image_to_string(gray_image, lang='spa')
            ocr_time = time.time() - ocr_start

            txt_output_path = Path(output_dir) / f"{pdf_path.stem}_p{page_num + 1}.txt"
            with open(txt_output_path, "w", encoding="utf-8") as f:
                f.write(text)

            lines = text.split('\n')
            print(f"   🔍 OCR completado en {ocr_time:.2f}s | {len(lines)} líneas")

            for idx, line in enumerate(lines):
                matchFactura = pat_fuzzy.search(line)
                if matchFactura:
                    num = regex.search(r'^\s*(\d{3,})', line)
                    if num:
                        # valor_factura = normalize_invoice(num.group(1))
                        valor_factura = (num.group(1))
                        print(f"   💼 Campo 51 (No. de factura): {valor_factura}")

                        valor_tasa = None
                        for next_line in lines[idx+1:]:
                            matchTasa = re.search(r'(?<!\d)(?:\d{1,3}(?:[.,]\d{3})+|\d+)[.,]\d{2}(?!\d)', next_line)
                            if matchTasa:
                                valor_tasa = normalize_number(matchTasa.group())
                                print(f"   💰 Campo 58 (Tasa de cambio $ cvs): {valor_tasa}")
                                break

                        if valor_factura and valor_tasa:
                            if not args.dry_run:
                                save_to_excel(valor_factura, valor_tasa, (page_num + 1, str(pdf_path.name)), excel_filename)
                            total_resultados += 1
                            print(f"   ✅ Datos encontrados: Factura {valor_factura}, Tasa {valor_tasa}")
                        else:
                            print(f"   ⚠️ No se encontró la tasa de cambio para la factura {valor_factura}")

            page_time = time.time() - page_start_time
            print(f"   ⏱️ Página procesada en {page_time:.2f} segundos")

    total_time = time.time() - start_time
    print("\n" + "=" * 60)
    print(f"🎉 PROCESAMIENTO COMPLETADO")
    print(f"📊 Archivo Excel: {excel_filename}")
    print(f"📈 Total de resultados encontrados: {total_resultados}")
    print(f"⏱️ Tiempo total de ejecución: {format_elapsed_time(total_time)}")
    print(f"🕐 Finalizado a las: {time.strftime('%H:%M:%S')}")
    print("=" * 60)

if __name__ == "__main__":
    main()
